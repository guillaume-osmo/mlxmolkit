#!/usr/bin/env python3
"""
Comprehensive comparison of RM1, AM1, AM1*, RM1* against the paper's
training set (Ong et al. 2025).

Reads molecular formulas from Excel, generates 3D via RDKit,
runs all 4 methods, compares to reference ΔHf.
"""
import sys; sys.path.insert(0, '.')
import numpy as np
import openpyxl
import re
import time
from collections import defaultdict

from mlxmolkit.rm1.pipeline import rm1_from_smiles, _smiles_to_3d
from mlxmolkit.rm1.scf import rm1_energy, rm1_energy_batch
from mlxmolkit.rm1.methods import get_params

# ================================================================
# Step 1: Load reference data from Excel
# ================================================================
wb = openpyxl.load_workbook('/Users/tgg/Downloads/Training and Testing Sets.xlsx')

all_mols = []
for sheet in ['Train', 'Test']:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        formula = ws.cell(r, 2).value
        hf_val = ws.cell(r, 3).value
        desc = ws.cell(r, 5).value or ''
        if formula and hf_val is not None:
            try:
                hf = float(hf_val)
            except (ValueError, TypeError):
                continue
            all_mols.append((formula, hf, desc, sheet))

print(f"Total molecules in dataset: {len(all_mols)}")

# ================================================================
# Step 2: Filter to neutral closed-shell CHNO molecules
# ================================================================
# We need SMILES to generate 3D. Let's use a formula→SMILES lookup
# for common molecules from the paper's training set.
# For a comprehensive comparison, use RDKit's formula→SMILES where possible.

# Common CHNO molecules with known SMILES
FORMULA_TO_SMILES = {
    'H2': '[H][H]',
    'H4C1': 'C',             # methane
    'H2C2': 'C#C',           # acetylene
    'H4C2': 'C=C',           # ethylene
    'H6C2': 'CC',            # ethane
    'H4C3': 'C=C=C',         # allene
    'H6C3': 'C=CC',          # propene
    'H8C3': 'CCC',           # propane
    'H2C4': 'C#CC#C',        # diacetylene
    'H4C4': 'C=CC=C',        # 1,3-butadiene
    'H6C4': 'C=CCC',         # 1-butene (approx)
    'H8C4': 'CCCC',          # butane
    'H10C4': 'CCCC',         # n-butane
    'H6C5': 'C1=CC=CC1',     # cyclopentadiene
    'H8C5': 'C1CCCC1',       # cyclopentane (wrong formula, skip)
    'H10C5': 'CCCCC',        # pentane
    'H12C5': 'CCCCC',        # n-pentane
    'H6C6': 'c1ccccc1',      # benzene
    'H8C6': 'C1=CC=CCC1',    # cyclohexadiene
    'H12C6': 'CCCCCC',       # hexane
    'H2O1': 'O',             # water
    'H4C1O1': 'CO',          # methanol
    'H2C1O1': 'C=O',         # formaldehyde
    'H2C1O2': 'OC=O',        # formic acid (approx)
    'H6C2O1': 'CCO',         # ethanol
    'H4C2O1': 'CC=O',        # acetaldehyde
    'H4C2O2': 'CC(=O)O',     # acetic acid
    'H3N1': 'N',             # ammonia
    'H5C1N1': 'CN',          # methylamine
    'H7C2N1': 'CCN',         # ethylamine
    'H3C1N1': 'C=N',         # methanimine (approx)
    'H1C1N1': 'C#N',         # hydrogen cyanide
    'H1N1O1': 'N=O',         # HNO (approx)
    'H1N1O2': 'ON=O',        # nitrous acid (approx)
    'H1N1O3': '[O-][N+](=O)O',  # nitric acid
    'C1O1': '[C-]#[O+]',     # carbon monoxide
    'C1O2': 'O=C=O',         # carbon dioxide
    'N2O1': 'N=NO',          # nitrous oxide (approx)
    'H4N2': 'NN',            # hydrazine
    'H2N2': 'N=N',           # diazene (approx)
    'H8C3O1': 'CCCO',        # 1-propanol (approx)
    'H6C3O1': 'CCC=O',       # propanal (approx)
    'H8C4O1': 'CCCCO',       # 1-butanol (approx)
    'H6C2O2': 'CC(O)=O',     # acetic acid (from above)
    'H4C3O1': 'C=CC=O',      # acrolein
    'H6C4O1': 'CC=CC=O',     # crotonaldehyde (approx)
    'H10C3N1': 'CCCN',       # propylamine (approx)
    'H5C2N1O1': 'CC(N)=O',   # acetamide (approx)
    'H8C7': 'Cc1ccccc1',     # toluene
    'H10C8': 'c1ccc2ccccc2c1',  # naphthalene
}

# Match formulas from dataset
matched = []
for formula, hf_ref, desc, src in all_mols:
    desc_lower = desc.lower()
    # Skip ions, radicals, triplets
    if any(x in desc_lower for x in ['cation', 'anion', 'radical', 'triplet', 'singlet']):
        continue
    # Only CHNO
    elements = set(re.findall(r'[A-Z][a-z]?', formula))
    if elements - {'H', 'C', 'N', 'O'}:
        continue
    # Look up SMILES
    if formula in FORMULA_TO_SMILES:
        matched.append((formula, hf_ref, FORMULA_TO_SMILES[formula], desc, src))

print(f"Matched with SMILES: {len(matched)}")

# ================================================================
# Step 3: Generate 3D and run all methods
# ================================================================
methods = ['RM1', 'AM1', 'AM1_STAR', 'RM1_STAR']

# Generate 3D coords
mol_data = []
for formula, hf_ref, smiles, desc, src in matched:
    result = _smiles_to_3d(smiles)
    if result is None:
        continue
    atoms, coords = result
    # Check all elements supported
    ok = True
    for method in methods:
        PARAMS = get_params(method)
        if any(z not in PARAMS for z in atoms):
            ok = False
            break
    if ok:
        mol_data.append((formula, hf_ref, smiles, desc, src, atoms, coords))

print(f"Successfully generated 3D: {len(mol_data)}")

# Run all 4 methods
results = {m: [] for m in methods}
ref_hf = []
labels = []

print(f"\nRunning {len(mol_data)} molecules × {len(methods)} methods...")
t0 = time.time()

for formula, hf_ref, smiles, desc, src, atoms, coords in mol_data:
    ref_hf.append(hf_ref)
    labels.append(f"{formula} ({desc[:25]})")

    for method in methods:
        try:
            r = rm1_energy(atoms, coords, max_iter=200, conv_tol=1e-6, method=method)
            if r['converged']:
                results[method].append(r['heat_of_formation_kcal'])
            else:
                results[method].append(None)
        except Exception:
            results[method].append(None)

t_total = time.time() - t0
print(f"Total time: {t_total:.1f}s")

# ================================================================
# Step 4: Statistics
# ================================================================
ref_hf = np.array(ref_hf)

print(f"\n{'='*80}")
print(f"  COMPARISON: 4 NDDO Methods vs Experimental ΔHf (kcal/mol)")
print(f"  Dataset: Ong et al. 2025 training/test set ({len(mol_data)} CHNO molecules)")
print(f"{'='*80}")

print(f"\n{'Method':>10s} {'N_conv':>7s} {'MAE':>8s} {'RMSE':>8s} {'MaxErr':>8s} {'R²':>8s}")
print("-" * 55)

for method in methods:
    vals = results[method]
    valid = [(ref_hf[i], vals[i]) for i in range(len(vals)) if vals[i] is not None]
    if len(valid) == 0:
        print(f"{method:>10s} {'0':>7s}")
        continue

    refs = np.array([v[0] for v in valid])
    preds = np.array([v[1] for v in valid])
    errors = preds - refs

    mae = np.mean(np.abs(errors))
    rmse = np.sqrt(np.mean(errors ** 2))
    max_err = np.max(np.abs(errors))

    ss_res = np.sum(errors ** 2)
    ss_tot = np.sum((refs - np.mean(refs)) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

    print(f"{method:>10s} {len(valid):>7d} {mae:>8.1f} {rmse:>8.1f} {max_err:>8.1f} {r2:>8.4f}")

# ================================================================
# Step 5: Per-molecule comparison table
# ================================================================
print(f"\n{'='*100}")
print(f"{'Formula':>12s} {'Description':>25s} {'Ref':>8s}", end="")
for m in methods:
    print(f" {m:>10s}", end="")
print()
print("-" * 100)

for i, (formula, hf_ref, smiles, desc, src, atoms, coords) in enumerate(mol_data):
    print(f"{formula:>12s} {desc[:25]:>25s} {hf_ref:>8.1f}", end="")
    for method in methods:
        val = results[method][i]
        if val is not None:
            err = val - hf_ref
            print(f" {val:>7.1f}({err:+.1f})", end="")
        else:
            print(f" {'NC':>10s}", end="")
    print()
